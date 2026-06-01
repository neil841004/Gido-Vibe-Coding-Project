#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表格結構分析器
分析遊戲數值表的結構、欄位和可能的關聯關係
"""

import os
import json
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
import re

class ExcelAnalyzer:
    def __init__(self, directory, output_directory=None):
        self.directory = Path(directory)
        self.output_directory = Path(output_directory) if output_directory else self.directory
        self.tables = {}
        self.relationships = []
        
    def analyze_all(self, progress_callback=None):
        """分析目錄中所有的 Excel 檔案

        progress_callback(done, total, filename) 會在每個檔案分析前被呼叫，
        供 GUI 顯示進度條。
        """
        print(f"開始分析 Excel 檔案 (目錄: {self.directory})...")

        files = sorted(self.directory.glob("*.xlsm"))
        total = len(files)
        for i, excel_file in enumerate(files):
            if progress_callback:
                progress_callback(i, total, excel_file.name)
            print(f"\n正在分析: {excel_file.name}")
            self.analyze_file(excel_file)
        if progress_callback:
            progress_callback(total, total, "")

        print("\n\n=== 分析完成 ===")
        self.detect_relationships()
        self.print_summary()
        self.save_results()
        
    def analyze_file(self, filepath):
        """分析單個 Excel 檔案"""
        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            file_key = filepath.stem  # 檔案名稱（不含副檔名）
            
            self.tables[file_key] = {
                'filename': filepath.name,
                'sheets': {}
            }
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 讀取表頭（假設在第一行）
                headers = []
                try:
                    first_row_gen = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                    first_row = next(first_row_gen, [])
                    headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(first_row)]
                except Exception:
                    headers = []
                
                # 統計資料行數
                row_count = max(0, sheet.max_row - 1)  # 扣除表頭
                
                self.tables[file_key]['sheets'][sheet_name] = {
                    'headers': headers,
                    'row_count': row_count,
                    'columns': {}
                }
                
                # 分析每個欄位的內容
                if headers:
                    for col_idx, header in enumerate(headers, start=1):
                        sample_values = []
                        for row in sheet.iter_rows(min_row=2, max_row=min(12, sheet.max_row), 
                                                  min_col=col_idx, max_col=col_idx, values_only=True):
                            if row and row[0] is not None:
                                sample_values.append(str(row[0]))
                        
                        self.tables[file_key]['sheets'][sheet_name]['columns'][header] = {
                            'samples': sample_values[:5]  # 只保存前5個樣本
                        }
                
                print(f"  - 工作表: {sheet_name} ({len(headers)} 欄位, {row_count} 筆資料)")
                
            wb.close()
            
        except Exception as e:
            print(f"  錯誤: {e}")
    
    def detect_relationships(self):
        """偵測表格之間的關聯關係"""
        print("\n\n=== 偵測關聯關係 ===")
        
        # 規則 1: 檔案名稱模式匹配（例如：Skill_Slicer 關聯到 Skill）
        for table_name in self.tables.keys():
            for other_table in self.tables.keys():
                if table_name != other_table:
                    # 檢查是否有下劃線分隔的命名模式
                    if '_' in table_name:
                        base_name = table_name.split('_')[0]
                        if base_name == other_table:
                            self.relationships.append({
                                'type': 'file_naming',
                                'from': other_table,
                                'to': table_name,
                                'rule': f'{other_table} -> {table_name} (命名規則)'
                            })
        
        # 規則 2: 欄位名稱包含其他表格名稱（例如：SkillID 可能關聯到 Skill 表）
        for table_name, table_data in self.tables.items():
            for sheet_name, sheet_data in table_data['sheets'].items():
                for header in sheet_data['headers']:
                    # 尋找包含其他表格名稱的欄位
                    for other_table in self.tables.keys():
                        if other_table != table_name:
                            # 檢查欄位名稱是否包含表格名稱
                            pattern = re.compile(rf'\b{other_table}\b', re.IGNORECASE)
                            if pattern.search(header):
                                self.relationships.append({
                                    'type': 'column_reference',
                                    'from': f'{table_name}.{sheet_name}',
                                    'to': other_table,
                                    'column': header,
                                    'rule': f'{table_name}.{sheet_name}.{header} -> {other_table}'
                                })
        
        # 規則 3: ID 欄位交叉引用
        id_columns = defaultdict(list)
        for table_name, table_data in self.tables.items():
            for sheet_name, sheet_data in table_data['sheets'].items():
                for header in sheet_data['headers']:
                    if 'ID' in header.upper():
                        id_columns[header].append(f'{table_name}.{sheet_name}')
        
        # 如果多個表格有相同的 ID 欄位名稱，它們可能有關聯
        for id_name, tables in id_columns.items():
            if len(tables) > 1:
                print(f"\n共用 ID 欄位 '{id_name}':")
                for t in tables:
                    print(f"  - {t}")
        
    def print_summary(self):
        """印出分析摘要"""
        print(f"\n\n=== Analysis Summary ===")
        print(f"Total analyzed: {len(self.tables)} Excel files")
        print(f"Total sheets: {sum(len(t['sheets']) for t in self.tables.values())}")
        print(f"Total relationships detected: {len(self.relationships)}")
        print("\nDetailed results saved to excel_structure.json")
    
    def save_results(self):
        """儲存分析結果為 JSON"""
        output = {
            'tables': self.tables,
            'relationships': self.relationships
        }
        
        output_file = self.output_directory / 'excel_structure.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        
        print(f"\n\n分析結果已儲存至: {output_file}")

def analyze_directory(input_path, output_path=None, progress_callback=None):
    """Wrapper for external calls"""
    input_dir = Path(input_path)

    # Use output_path if provided, else define relative to input
    if output_path:
        out_dir = Path(output_path)
    else:
        out_dir = input_dir

    output_file = out_dir / 'excel_structure.json'

    print(f"Analyzing directory: {input_dir}")
    print(f"Output Structure File: {output_file}")

    analyzer = ExcelAnalyzer(input_dir, out_dir)
    analyzer.analyze_all(progress_callback=progress_callback)

def main():
    import sys
    
    current_dir = Path(__file__).parent
    input_dir = current_dir
    
    if len(sys.argv) > 1:
        input_dir = Path(sys.argv[1])
        
    analyze_directory(input_dir, current_dir)

if __name__ == '__main__':
    main()
