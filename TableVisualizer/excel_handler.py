#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 處理器 (Cross-Platform)
負責：
1. 開啟特定工作表 (Windows: COM, Mac: subprocess open)
2. 讀取資料 (OpenPyXL)
3. 寫入資料 (Windows: COM, Mac: OpenPyXL with keep_vba=True)
"""

import os
import sys
import subprocess
from pathlib import Path
import openpyxl

IS_WINDOWS = sys.platform.startswith('win')

if IS_WINDOWS:
    import win32com.client
    import pythoncom

class ExcelHandler:
    def __init__(self, base_dir):
        self.base_dir = Path(base_dir)

    def open_sheet(self, filename, sheet_name):
        """
        開啟 Excel 並跳轉到指定工作表
        """
        file_path = self.base_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        abs_path = str(file_path.absolute())

        if IS_WINDOWS:
            try:
                pythoncom.CoInitialize()
                try:
                    excel = win32com.client.GetActiveObject("Excel.Application")
                except:
                    excel = win32com.client.Dispatch("Excel.Application")
                
                excel.Visible = True
                
                wb = None
                
                for w in excel.Workbooks:
                    if w.FullName.lower() == abs_path.lower():
                        wb = w
                        break
                
                if wb is None:
                    wb = excel.Workbooks.Open(abs_path)
                    
                try:
                    ws = wb.Worksheets(sheet_name)
                    ws.Activate()
                except Exception as e:
                    print(f"Failed to activate sheet {sheet_name}: {e}")
                    wb.Worksheets(1).Activate()
                    
            except Exception as e:
                print(f"COM Error: {e}")
                os.startfile(file_path)
            finally:
                pythoncom.CoUninitialize()
        else:
            # macOS / Linux
            try:
                if sys.platform == 'darwin':
                    # Use AppleScript to open and select sheet
                    # We open via AppleScript to ensure it's ready before acting on it
                    safe_path = abs_path.replace('"', '\\"')
                    safe_sheetname = sheet_name.replace('"', '\\"')
                    
                    script = f'''
                    tell application "Microsoft Excel"
                        activate
                        open (POSIX file "{safe_path}")
                        
                        try
                            activate object worksheet "{safe_sheetname}" of active workbook
                        on error
                            -- iterate to find if active workbook assumption failed
                            repeat with wb in workbooks
                                try
                                    activate object worksheet "{safe_sheetname}" of wb
                                    exit repeat
                                end try
                            end repeat
                        end try
                    end tell
                    '''
                    
                    subprocess.run(['osascript', '-e', script], capture_output=True)
                    
                else:
                    subprocess.call(['xdg-open', abs_path])
                
                print(f"Opened {filename} -> {sheet_name}")
            except Exception as e:
                print(f"Failed to open file: {e}")

    def get_sheet_data(self, filename, sheet_name, param_row_idx=2):
        """
        使用 OpenPyXL 讀取資料 (唯讀模式，安全)
        """
        file_path = self.base_dir / filename
        if not file_path.exists(): return []

        try:
            # data_only=True 讀取計算後的值
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return []
                
            ws = wb[sheet_name]
            
            headers = []
            values = []
            
            # 讀取第一列
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
                
            # 讀取指定列 (參數值)
            if ws.max_row >= param_row_idx:
                for cell in ws[param_row_idx]:
                    val = cell.value
                    values.append(str(val) if val is not None else "")
            else:
                values = [""] * len(headers)
                
            wb.close()
            
            data = []
            for i, header in enumerate(headers):
                if not header: continue
                val = values[i] if i < len(values) else ""
                data.append({'header': header, 'value': val, 'col_idx': i + 1})
                
            return data
            
        except Exception as e:
            print(f"Read Error: {e}")
            return []

    def get_all_sheet_data(self, filename, sheet_name):
        """
        Retrieves all data from a sheet as a list of lists.
        Returns: (headers, data_rows) where data_rows is list of lists
        """
        file_path = self.base_dir / filename
        if not file_path.exists(): return [], []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return [], []
            ws = wb[sheet_name]
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert None to "" and ensures strings
                r_data = [str(cell) if cell is not None else "" for cell in row]
                rows.append(r_data)
            
            wb.close()
            
            if not rows: return [], []
            
            headers = rows[0]
            data = rows[1:]
            return headers, data
            
        except Exception as e:
            print(f"Read All Error: {e}")
            return [], []

    def update_cell_value(self, filename, sheet_name, col_idx, row_idx, value):
        """
        寫入資料
        Windows: 使用 COM (防止 .xlsm 損壞)
        Mac/Others: 使用 OpenPyXL (keep_vba=True)
        """
        file_path = self.base_dir / filename
        if not file_path.exists():
            return False, "File not found"
            
        # Value conversion logic shared
        def convert_value(val):
            if val.lower() == 'true':
                return True
            elif val.lower() == 'false':
                return False
            else:
                try:
                    f_val = float(val)
                    if f_val.is_integer():
                        return int(f_val)
                    else:
                        return f_val
                except ValueError:
                    return val

        converted_value = convert_value(value)

        if IS_WINDOWS:
            try:
                pythoncom.CoInitialize()
                try:
                    excel = win32com.client.GetActiveObject("Excel.Application")
                except:
                    excel = win32com.client.Dispatch("Excel.Application")
                
                wb = None
                abs_path = str(file_path.absolute())
                is_opened_by_me = False
                
                for w in excel.Workbooks:
                    if w.FullName.lower() == abs_path.lower():
                        wb = w
                        break
                
                if wb is None:
                    wb = excel.Workbooks.Open(abs_path)
                    is_opened_by_me = True
                    
                try:
                    ws = wb.Worksheets(sheet_name)
                    cell = ws.Cells(row_idx, col_idx)
                    cell.Value = converted_value
                    wb.Save()
                    
                    if is_opened_by_me:
                        wb.Close()
                        
                except Exception as e:
                    return False, f"Write Error: {e}"
                    
                return True, "Success"
                
            except Exception as e:
                return False, f"COM Error: {e}"
            finally:
                pythoncom.CoUninitialize()

        elif sys.platform == 'darwin':
            # macOS implementation using AppleScript (Safe for complex .xlsm)
            try:
                abs_path = str(file_path.absolute())
                safe_path = abs_path.replace('"', '\\"')
                safe_sheetname = sheet_name.replace('"', '\\"')
                
                # Format value for AppleScript
                val_repr = ""
                if isinstance(converted_value, bool):
                    val_repr = "true" if converted_value else "false"
                elif isinstance(converted_value, (int, float)):
                    val_repr = str(converted_value)
                else:
                    # Escape double quotes and backslashes for AppleScript
                    # AppleScript string escaping is tricky, usually backslash escapes double quote
                    safe_val_str = str(converted_value).replace('\\', '\\\\').replace('"', '\\"')
                    val_repr = f'"{safe_val_str}"'

                script = f'''
                tell application "Microsoft Excel"
                    set targetWorkbook to open workbook workbook file name "{safe_path}"
                    
                    tell worksheet "{safe_sheetname}" of targetWorkbook
                        set value of cell {col_idx} of row {row_idx} to {val_repr}
                    end tell
                    
                    save targetWorkbook
                end tell
                '''
                
                result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
                
                if result.returncode != 0:
                    return False, f"AppleScript Error: {result.stderr}"
                    
                return True, "Success"
                
            except Exception as e:
                return False, f"Write Error (AppleScript): {e}"

        else:
            # Linux / Fallback implementation using OpenPyXL
            try:
                # keep_vba=True is crucial for .xlsm
                wb = openpyxl.load_workbook(file_path, keep_vba=True)
                if sheet_name not in wb.sheetnames:
                    return False, f"Sheet '{sheet_name}' not found"
                
                ws = wb[sheet_name]
                # OpenPyXL uses 1-based indexing
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = converted_value
                
                wb.save(file_path)
                return True, "Success"
            except Exception as e:
                return False, f"Write Error (OpenPyXL): {e}"
